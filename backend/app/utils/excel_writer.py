from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import cast

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

HEADER = [
    "序号",
    "识别时间",
    "物料编码",
    "数量",
    "批次",
    "日期",
    "品牌",
    "电气特性",
    "原始OCR文本",
    "图片文件名",
]


def initialize_excel(path: Path) -> None:
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "识别记录"
    ws.append(HEADER)
    wb.save(path)


def append_result(path: Path, result) -> None:
    if not path.exists():
        initialize_excel(path)

    wb = load_workbook(path)
    ws = cast(Worksheet, wb.active)
    next_index = ws.max_row
    ws.append(
        [
            next_index,
            result.recognized_at.strftime("%Y-%m-%d %H:%M:%S"),
            result.material_code,
            result.quantity,
            result.batch,
            result.date,
            result.brand,
            result.electrical_characteristics,
            result.raw_ocr_text,
            result.image_filename,
        ]
    )
    wb.save(path)