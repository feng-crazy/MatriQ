"""测试 Excel 写入工具"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.models.pipeline import RecognitionResult
from app.utils.excel_writer import append_result, initialize_excel


class TestExcelWriter:
    """测试 Excel 写入功能"""

    def test_initialize_excel(self, temp_data_dir: Path):
        """测试初始化 Excel 文件"""
        excel_path = temp_data_dir / "test.xlsx"
        initialize_excel(excel_path)

        assert excel_path.exists()

        # 验证表头
        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 1
        assert ws.cell(1, 1).value == "序号"
        assert ws.cell(1, 2).value == "识别时间"
        assert ws.cell(1, 3).value == "物料编码"
        assert ws.cell(1, 10).value == "图片文件名"

    def test_append_result(self, temp_data_dir: Path):
        """测试追加结果"""
        excel_path = temp_data_dir / "test.xlsx"
        initialize_excel(excel_path)

        # 创建模拟结果对象
        result = RecognitionResult(
            id=1,
            pipeline_id=1,
            material_code="SL-IND-1008-100",
            quantity=4000,
            batch="B2511A",
            date="2025-11-30",
            brand="Sunlord",
            electrical_characteristics="10uH ±10%",
            raw_ocr_text="Sunlord SL-IND-1008-100 Qty:4000",
            image_filename="test.jpg",
            recognized_at=datetime(2025, 11, 30, 12, 0, 0),
        )

        append_result(excel_path, result)

        # 验证数据
        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 2  # 表头 + 1 行数据

        # 验证数据行
        assert ws.cell(2, 1).value == 1  # 序号
        assert ws.cell(2, 2).value == "2025-11-30 12:00:00"  # 识别时间
        assert ws.cell(2, 3).value == "SL-IND-1008-100"  # 物料编码
        assert ws.cell(2, 4).value == 4000  # 数量
        assert ws.cell(2, 5).value == "B2511A"  # 批次
        assert ws.cell(2, 6).value == "2025-11-30"  # 日期
        assert ws.cell(2, 7).value == "Sunlord"  # 品牌
        assert ws.cell(2, 8).value == "10uH ±10%"  # 电气特性
        assert ws.cell(2, 9).value == "Sunlord SL-IND-1008-100 Qty:4000"  # 原始OCR文本
        assert ws.cell(2, 10).value == "test.jpg"  # 图片文件名

    def test_append_multiple_results(self, temp_data_dir: Path):
        """测试追加多条结果"""
        excel_path = temp_data_dir / "test.xlsx"
        initialize_excel(excel_path)

        # 追加第一条
        result1 = RecognitionResult(
            id=1,
            pipeline_id=1,
            material_code="SL-IND-1008-100",
            quantity=4000,
            recognized_at=datetime(2025, 11, 30, 12, 0, 0),
        )
        append_result(excel_path, result1)

        # 追加第二条
        result2 = RecognitionResult(
            id=2,
            pipeline_id=1,
            material_code="SL-IND-1009-200",
            quantity=5000,
            recognized_at=datetime(2025, 11, 30, 13, 0, 0),
        )
        append_result(excel_path, result2)

        # 验证
        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 3  # 表头 + 2 行数据

        assert ws.cell(2, 1).value == 1
        assert ws.cell(3, 1).value == 2

    def test_append_to_nonexistent_file(self, temp_data_dir: Path):
        """测试追加到不存在的文件（应自动创建）"""
        excel_path = temp_data_dir / "new.xlsx"

        result = RecognitionResult(
            id=1,
            pipeline_id=1,
            material_code="SL-IND-1008-100",
            recognized_at=datetime(2025, 11, 30, 12, 0, 0),
        )

        append_result(excel_path, result)

        # 文件应该被创建
        assert excel_path.exists()

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 2  # 表头 + 1 行数据

